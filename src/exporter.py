"""
src/exporter.py - Excel 3-Sheet 报表导出与专业趋势图表绘制
"""
import os
from typing import Dict, List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from src.utils import setup_logger, get_beijing_now

logger = setup_logger("Exporter")

class ExcelExporter:
    def __init__(self, output_dir: str = "./reports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_daily_report(
        self,
        site_a_df: pd.DataFrame,
        site_a_meta: Dict,
        site_b_latest_df: pd.DataFrame,
        site_b_trend_df: pd.DataFrame,
        period_headers: List[str],
        site_b_meta: Dict
    ) -> str:
        """
        生成命名为 Daily_Report_YYYYMMDD.xlsx 的 3-Sheet 专业分析报表
        Sheet 1: 生意社大宗商品监测
        Sheet 2: 统计局重要生产资料价格(最新)
        Sheet 3: 生产资料价格走势(近10次) [含各细项趋势图表]
        """
        today_str = get_beijing_now().strftime("%Y%m%d")
        file_name = f"Daily_Report_{today_str}.xlsx"
        file_path = os.path.join(self.output_dir, file_name)
        
        logger.info(f"准备生成 3-Sheet 每日 Excel 分析报表: {file_path}")

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Sheet 1: 站点 A 数据
            df_a = site_a_df if not site_a_df.empty else pd.DataFrame({"提示": ["今日未获取到有效数据或页面正在安全检查"]})
            df_a.to_excel(writer, sheet_name="生意社大宗商品监测", index=False, startrow=3)

            # Sheet 2: 站点 B 最新一期
            df_b = site_b_latest_df if not site_b_latest_df.empty else pd.DataFrame({"提示": ["统计局暂无新一期流通领域生产资料价格发布"]})
            df_b.to_excel(writer, sheet_name="统计局生产资料(最新期)", index=False, startrow=3)

            # Sheet 3: 过去 10 次价格走势矩阵
            df_trend = site_b_trend_df if not site_b_trend_df.empty else pd.DataFrame({"提示": ["未获取到过去多期历史数据"]})
            df_trend.to_excel(writer, sheet_name="生产资料走势(近10次)", index=False, startrow=3)

        # 进行单元格美化排版与原生趋势图表绘制
        self._apply_styling_and_charts(file_path, site_a_meta, site_b_meta, period_headers, site_b_trend_df)
        
        logger.info(f"🎉 3-Sheet 报表生成完毕，包含原生趋势图: {file_path}")
        return file_path

    def _apply_styling_and_charts(
        self,
        file_path: str,
        meta_a: Dict,
        meta_b: Dict,
        period_headers: List[str],
        trend_df: pd.DataFrame
    ):
        wb = load_workbook(file_path)

        # 核心样式定义
        title_font = Font(name="微软雅黑", size=13, bold=True, color="1F4E78")
        meta_font = Font(name="微软雅黑", size=9, italic=True, color="595959")
        header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        data_font = Font(name="微软雅黑", size=9.5)
        alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        up_font = Font(name="微软雅黑", size=9.5, color="C00000", bold=True)
        down_font = Font(name="微软雅黑", size=9.5, color="008000", bold=True)

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        now_str = get_beijing_now().strftime("%Y-%m-%d %H:%M:%S")

        # -------------------------------------------------------------
        # 1. 格式化 Sheet 1 & Sheet 2
        # -------------------------------------------------------------
        sheet_configs = [
            ("生意社大宗商品监测", meta_a, "生意社 (100ppi) 大宗商品每日价格监测报表"),
            ("统计局生产资料(最新期)", meta_b, f"国家统计局 - {meta_b.get('latest_title', '流通领域重要生产资料市场价格变动情况')}")
        ]

        for sheet_name, meta, report_title in sheet_configs:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            ws.views.sheetView[0].showGridLines = True

            ws.cell(row=1, column=1, value=report_title).font = title_font
            meta_desc = f"【数据来源】: {meta.get('url')}  |  【生成时间】: {now_str}  |  【记录数】: {meta.get('row_count', len(ws['A'])-4)} 条"
            ws.cell(row=2, column=1, value=meta_desc).font = meta_font

            header_row = 4
            max_row = ws.max_row
            max_col = ws.max_column

            for col in range(1, max_col + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[header_row].height = 28

            # 识别关键数据列索引（本期价格、涨跌额、涨跌幅）
            col_diff_idx = None
            col_rate_idx = None
            for c in range(1, max_col + 1):
                h = str(ws.cell(row=header_row, column=c).value or "")
                if "涨跌(" in h or "比上期" in h or "价格涨跌" in h:
                    col_diff_idx = c
                elif "涨跌幅" in h or "涨跌" in h:
                    col_rate_idx = c

            for row in range(header_row + 1, max_row + 1):
                ws.row_dimensions[row].height = 20
                is_alt = (row % 2 == 0)

                # 判断该行的涨跌状态 (1: 涨/红, -1: 跌/绿, 0: 平)
                trend_status = 0
                for c_idx in [col_diff_idx, col_rate_idx]:
                    if c_idx:
                        try:
                            v_raw = str(ws.cell(row=row, column=c_idx).value or "").replace(",", "").replace("%", "").strip()
                            v_num = float(v_raw)
                            if v_num > 0:
                                trend_status = 1
                                break
                            elif v_num < 0:
                                trend_status = -1
                                break
                        except Exception:
                            pass

                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    if is_alt:
                        cell.fill = alt_fill

                    col_header = str(ws.cell(row=header_row, column=col).value or "")
                    val_str = str(cell.value or "")

                    # 核心需求：本期价格、涨跌额、涨跌幅 按照 A 股规则「涨标红、跌标绿」
                    if any(k in col_header for k in ["本期价格", "现价", "价格(元)"]):
                        if trend_status == 1:
                            cell.font = up_font
                        elif trend_status == -1:
                            cell.font = down_font
                    elif any(k in col_header for k in ["涨跌", "比上期"]):
                        try:
                            v_val = float(val_str.replace(",", "").replace("%", "").strip())
                            if v_val > 0:
                                cell.font = up_font
                            elif v_val < 0:
                                cell.font = down_font
                        except Exception:
                            if trend_status == 1:
                                cell.font = up_font
                            elif trend_status == -1:
                                cell.font = down_font
                    
                    # 对齐策略：文本严格左对齐；单位居中；所有价格/涨跌/涨幅/数值列（含 0 和 0.0）一律严格右对齐
                    if any(k in col_header for k in ["产品", "商品", "名称", "规格", "型号", "分类", "大类", "提示", "来源"]):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif "单位" in col_header:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif any(k in col_header for k in ["涨", "跌", "幅", "价", "额", "比", "旬", "日", "月", "期", "变动"]) or self._is_numeric(val_str):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # 列宽自适应
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = 0
                for cell in col:
                    if cell.row < 3:
                        continue
                    if cell.value:
                        val_str = str(cell.value)
                        width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                        max_len = max(max_len, width)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # -------------------------------------------------------------
        # 2. 格式化 Sheet 3 并绘制趋势图表 (LineChart)
        # -------------------------------------------------------------
        trend_sheet_name = "生产资料走势(近10次)"
        if trend_sheet_name in wb.sheetnames and not trend_df.empty:
            ws3 = wb[trend_sheet_name]
            ws3.views.sheetView[0].showGridLines = True

            report_title_3 = f"国家统计局 - 流通领域50种重要生产资料市场价格过去 {len(period_headers)} 次变动与趋势图"
            ws3.cell(row=1, column=1, value=report_title_3).font = title_font
            meta_desc_3 = f"【历史覆盖期数】: {len(period_headers)} 期 ({period_headers[0]} 至 {period_headers[-1]})  |  【生成时间】: {now_str}"
            ws3.cell(row=2, column=1, value=meta_desc_3).font = meta_font

            header_row = 4
            max_row = ws3.max_row
            max_col = ws3.max_column

            # 表头美化
            for col in range(1, max_col + 1):
                cell = ws3.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws3.row_dimensions[header_row].height = 28

            # 数据单元格美化
            for row in range(header_row + 1, max_row + 1):
                ws3.row_dimensions[row].height = 20
                is_alt = (row % 2 == 0)
                for col in range(1, max_col + 1):
                    cell = ws3.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    if is_alt:
                        cell.fill = alt_fill

                    col_header = str(ws3.cell(row=header_row, column=col).value or "")
                    val_str = str(cell.value or "")

                    # 涨跌幅特殊标红/标绿
                    if "涨跌" in col_header:
                        try:
                            val_num = float(cell.value)
                            if val_num > 0:
                                cell.font = up_font
                            elif val_num < 0:
                                cell.font = down_font
                        except Exception:
                            pass

                    # 对齐策略：文本严格左对齐；单位居中；所有价格/涨跌/涨幅/数值列（含 0 和 0.0）一律严格右对齐
                    if any(k in col_header for k in ["产品", "商品", "名称", "规格", "型号", "分类", "大类", "提示", "来源"]):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif "单位" in col_header:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif any(k in col_header for k in ["涨", "跌", "幅", "价", "额", "比", "旬", "日", "月", "期", "变动"]) or self._is_numeric(val_str):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # 列宽计算
            for col in ws3.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = 0
                for cell in col:
                    if cell.row < 3:
                        continue
                    if cell.value:
                        val_str = str(cell.value)
                        width = sum(2 if ord(c) > 127 else 1 for c in val_str)
                        max_len = max(max_len, width)
                ws3.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # -------------------------------------------------------------
            # 3. 绘制每个分类大类及重点商品的趋势折线图 (Native Excel Charts)
            # -------------------------------------------------------------
            self._create_trend_charts(ws3, trend_df, period_headers, header_row=4, data_start_row=5)

        wb.save(file_path)

    def _create_trend_charts(
        self,
        ws,
        trend_df: pd.DataFrame,
        period_headers: List[str],
        header_row: int = 4,
        data_start_row: int = 5
    ):
        """
        在 Sheet 3 下方插入各分类及细项的原生折线趋势图
        """
        if not period_headers:
            return

        col_product = 2
        col_period_start = 5
        col_period_end = 5 + len(period_headers) - 1

        # X 轴期数引用 (第 4 行的期数表头)
        x_values = Reference(ws, min_col=col_period_start, max_col=col_period_end, min_row=header_row, max_row=header_row)

        # 获取所有分类大类及其对应的行区间
        categories = []
        current_cat = None
        start_r = data_start_row

        for i, row in trend_df.iterrows():
            cat = str(row["分类大类"])
            current_row_num = data_start_row + i
            if cat != current_cat:
                if current_cat is not None:
                    categories.append((current_cat, start_r, current_row_num - 1))
                current_cat = cat
                start_r = current_row_num
        if current_cat is not None:
            categories.append((current_cat, start_r, data_start_row + len(trend_df) - 1))

        # 图表起始放置位置：在数据表格下方（留出 3 行空隙）
        chart_start_row = data_start_row + len(trend_df) + 3

        logger.info(f"正在生成 {len(categories)} 张大类细项价格走势折线图...")

        # 采用双列并排布局放置图表
        chart_cols = ["B", "N"]
        chart_height_rows = 18

        # 高对比度专业配色库 (D3/Tableau 经典高区分度色系，杜绝单一色系混淆)
        DISTINCT_COLORS = [
            "1F77B4",  # 经典深蓝
            "FF7F0E",  # 活力鲜橙
            "2CA02C",  # 森林翠绿
            "D62728",  # 绯红明亮
            "9467BD",  # 典雅高贵紫
            "8C564B",  # 暖调棕褐
            "E377C2",  # 亮丽玫粉
            "17BECF",  # 科技青天蓝
            "BCBD22",  # 橄榄金黄
            "393B79",  # 深邃藏青
            "E6550D",  # 琥珀深橙
            "756BB1",  # 薰衣草紫
        ]

        for idx, (cat_name, min_r, max_r) in enumerate(categories):
            chart = LineChart()
            chart.title = f"{cat_name} - 细项价格过去 {len(period_headers)} 期走势 (单位: 元)"
            chart.width = 19
            chart.height = 10

            # 明确指定 X 轴置于底部(bottom)、Y 轴置于左侧(left)
            chart.x_axis.axPos = "b"
            chart.y_axis.axPos = "l"
            chart.x_axis.tickLblPos = "low"
            chart.y_axis.tickLblPos = "low"
            chart.x_axis.crosses = "autoZero"
            chart.y_axis.crosses = "autoZero"
            chart.x_axis.title = "发布期数"
            chart.y_axis.title = "价格 (元)"
            if chart.legend:
                chart.legend.legendPos = "r"

            # 数据引用：该分类下所有商品行的价格数据
            data = Reference(ws, min_col=col_period_start, max_col=col_period_end, min_row=min_r, max_row=max_r)
            chart.add_data(data, from_rows=True, titles_from_data=False)
            chart.set_categories(x_values)

            # 自定义每条折线的独立高区分度色彩与精致细线条
            for s_idx, series in enumerate(chart.series):
                series_row = min_r + s_idx
                product_name = str(ws.cell(row=series_row, column=col_product).value or f"商品{s_idx+1}")
                series.tx = SeriesLabel(v=product_name)

                # 分配高对比度独立颜色
                color = DISTINCT_COLORS[s_idx % len(DISTINCT_COLORS)]

                # 设置精致细线条 (15000 EMU ≈ 1.2pt)
                series.graphicalProperties.line.solidFill = color
                series.graphicalProperties.line.width = 15000

                # 标记点设小，颜色与线条一致
                series.marker.symbol = "circle"
                series.marker.size = 3
                series.marker.graphicalProperties.solidFill = color
                series.marker.graphicalProperties.line.solidFill = color
                series.smooth = True

            # 计算图表摆放单元格 (双列布局)
            row_offset = (idx // 2) * chart_height_rows
            col_target = chart_cols[idx % 2]
            cell_loc = f"{col_target}{chart_start_row + row_offset}"

            ws.add_chart(chart, cell_loc)

    def _is_numeric(self, val_str: str) -> bool:
        try:
            float(val_str.replace(",", "").replace("%", ""))
            return True
        except Exception:
            return False
